import streamlit as st
import pandas as pd
import requests
import time
import re
import io
import json
import zipfile
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="SEC Financial Data Extractor", layout="wide")

HEADERS = {"User-Agent": "FinancialDataExtractor contact@example.com"}

# ── EDGAR helpers ──────────────────────────────────────────────────────────────

def get_cik(ticker: str) -> str | None:
    r = requests.get("https://www.sec.gov/files/company_tickers.json", headers=HEADERS, timeout=15)
    for item in r.json().values():
        if item["ticker"].upper() == ticker.upper():
            return str(item["cik_str"]).zfill(10)
    return None

def get_company_name(cik: str) -> str:
    r = requests.get(f"https://data.sec.gov/submissions/CIK{cik}.json", headers=HEADERS, timeout=15)
    return r.json().get("name", "Unknown")

def get_xbrl_facts(cik: str) -> dict:
    r = requests.get(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json", headers=HEADERS, timeout=30)
    return r.json().get("facts", {})

def get_filings_index(cik: str, form_types: list[str]) -> list[dict]:
    r = requests.get(f"https://data.sec.gov/submissions/CIK{cik}.json", headers=HEADERS, timeout=15)
    data = r.json()
    filings = data.get("filings", {}).get("recent", {})
    forms   = filings.get("form", [])
    dates   = filings.get("filingDate", [])
    accs    = filings.get("accessionNumber", [])
    docs    = filings.get("primaryDocument", [])
    results = []
    for i, form in enumerate(forms):
        if form in form_types:
            results.append({
                "form": form,
                "date": dates[i],
                "accession": accs[i].replace("-", ""),
                "accession_fmt": accs[i],
                "primary_doc": docs[i],
            })
    return results


# ── EDGAR R-files: get the actual statement structure from filings ─────────────
#
# Every EDGAR filing has an index page that lists all documents, including
# structured XBRL "R" viewer files. These tell us exactly which concepts
# belong on which statement (IS vs BS vs CF) and in what order, as reported.

def get_filing_index_page(cik: str, accession: str) -> dict:
    """Fetch the filing index JSON to find all documents."""
    url = f"https://data.sec.gov/submissions/CIK{cik}.json"
    # Use the EDGAR filing index endpoint
    acc_fmt = f"{accession[:10]}-{accession[10:12]}-{accession[12:]}"
    url = f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={cik}&type=10-K&dateb=&owner=include&count=5"
    index_url = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{accession}/index.json"
    try:
        r = requests.get(index_url, headers=HEADERS, timeout=15)
        if r.status_code == 200:
            return r.json()
    except Exception:
        pass
    return {}

def get_r_files(cik: str, accession: str) -> dict[str, str]:
    """
    Fetch the R-viewer JSON files from an EDGAR filing.
    Returns {statement_type: url} for IS, BS, CF statements.
    """
    index_url = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{accession}/index.json"
    try:
        r = requests.get(index_url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            return {}
        files = r.json().get("directory", {}).get("item", [])
        # Look for R files (XBRL viewer) — they are named R1.htm, R2.htm, etc.
        # or the main XBRL instance document
        r_files = {f["name"]: f["name"] for f in files if re.match(r"R\d+\.htm", f.get("name", ""))}
        return r_files
    except Exception:
        return {}


def fetch_xbrl_viewer_statements(cik: str, accession: str) -> dict[str, list[dict]]:
    """
    Use EDGAR's XBRL inline viewer API to get structured statement data.
    Returns {statement_name: [{label, concept, value, level, is_total}]}
    """
    # The EDGAR viewer API endpoint
    acc_dashes = f"{accession[:10]}-{accession[10:12]}-{accession[12:]}"
    base = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{accession}/"

    # Get filing index to find the XBRL instance document
    index_url = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{accession}/index.json"
    try:
        r = requests.get(index_url, headers=HEADERS, timeout=15)
        if r.status_code != 200:
            return {}
        items = r.json().get("directory", {}).get("item", [])
    except Exception:
        return {}

    # Find the XBRL calculation or presentation linkbase, or R htm files
    r_files = sorted([f["name"] for f in items if re.match(r"R\d+\.htm", f.get("name", ""))],
                     key=lambda x: int(re.search(r"\d+", x).group()))

    if not r_files:
        return {}

    statements = {}
    statement_keywords = {
        "IS":  ["income", "operation", "earnings", "profit", "loss", "comprehensive"],
        "BS":  ["balance", "financial position", "asset", "liabilit"],
        "CF":  ["cash flow", "cash and cash"],
    }

    for fname in r_files[:40]:  # check first 40 R files
        url = base + fname
        try:
            r = requests.get(url, headers=HEADERS, timeout=10)
            if r.status_code != 200:
                continue
            soup = BeautifulSoup(r.text, "html.parser")

            # Get the statement title
            title_el = soup.find("th", class_="pl") or soup.find("div", class_="rh") or soup.find("th")
            title_text = title_el.get_text(strip=True).lower() if title_el else ""

            # Classify statement type
            stmt_type = None
            for stype, kws in statement_keywords.items():
                if any(kw in title_text for kw in kws):
                    stmt_type = stype
                    break
            if not stmt_type:
                continue

            # Don't overwrite if we already have a better version of this statement
            if stmt_type in statements:
                continue

            # Parse the table
            rows = []
            for tr in soup.find_all("tr"):
                tds = tr.find_all(["td", "th"])
                if len(tds) < 2:
                    continue
                label_el = tds[0]
                label = label_el.get_text(strip=True)
                if not label:
                    continue

                # Get concept name from onclick or data attributes
                concept = ""
                onclick = label_el.get("onclick", "") or tr.get("onclick", "")
                m = re.search(r"'([A-Za-z]+)'", onclick)
                if m:
                    concept = m.group(1)

                # Get indentation level
                style = label_el.get("style", "")
                indent_m = re.search(r"padding-left:\s*(\d+)", style)
                level = int(indent_m.group(1)) // 10 if indent_m else 0

                # Get values (skip first col which is label)
                values = [td.get_text(strip=True).replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
                          for td in tds[1:]]

                rows.append({
                    "label": label,
                    "concept": concept,
                    "level": level,
                    "values": values,
                    "is_total": "total" in label.lower() or level == 0,
                })

            if rows:
                statements[stmt_type] = {"title": title_text, "rows": rows, "file": fname}

        except Exception:
            continue
        time.sleep(0.1)

    return statements


# ── Primary approach: XBRL companyfacts API with smart period filtering ────────
#
# This is the most reliable approach. We get all facts, then for each filing
# period we look up which concepts were actually on each statement by checking
# the company's R-files for one representative filing, then use that concept
# list to pull time-series data from companyfacts.

# Statement concept assignments — seeded from common GAAP taxonomy,
# supplemented by R-file discovery per company.

# Hard classification for unambiguous concepts
CONCEPT_STATEMENT = {
    # Income Statement
    "Revenues": "IS", "RevenueFromContractWithCustomerExcludingAssessedTax": "IS",
    "RevenueFromContractWithCustomerIncludingAssessedTax": "IS", "SalesRevenueNet": "IS",
    "SalesRevenueGoodsNet": "IS", "SalesRevenueServicesNet": "IS", "RevenueNet": "IS",
    "RoyaltyRevenue": "IS", "LicenseAndServicesRevenue": "IS", "LicensesRevenue": "IS",
    "CostOfRevenue": "IS", "CostOfGoodsSold": "IS", "CostOfGoodsAndServicesSold": "IS",
    "CostOfServices": "IS", "GrossProfit": "IS",
    "ResearchAndDevelopmentExpense": "IS", "ResearchAndDevelopmentExpenseExcludingAcquiredInProcessCost": "IS",
    "SellingGeneralAndAdministrativeExpense": "IS", "GeneralAndAdministrativeExpense": "IS",
    "SellingAndMarketingExpense": "IS", "SellingExpense": "IS",
    "AmortizationOfIntangibleAssets": "IS", "AmortizationOfAcquiredIntangibleAssets": "IS",
    "GoodwillImpairmentLoss": "IS", "ImpairmentOfIntangibleAssetsExcludingGoodwill": "IS",
    "RestructuringCharges": "IS", "RestructuringAndRelatedCostIncurredCost": "IS",
    "BusinessCombinationAcquisitionRelatedCosts": "IS",
    "OperatingExpenses": "IS", "CostsAndExpenses": "IS",
    "OperatingIncomeLoss": "IS",
    "InterestExpense": "IS", "InterestAndDebtExpense": "IS", "InterestExpenseDebt": "IS",
    "InterestIncomeExpenseNet": "IS", "InterestIncomeExpenseNonoperatingNet": "IS",
    "InvestmentIncomeNonoperating": "IS", "InvestmentIncomeInterest": "IS",
    "NonoperatingIncomeExpense": "IS", "OtherNonoperatingIncomeExpense": "IS",
    "OtherNonoperatingIncome": "IS", "OtherNonoperatingExpense": "IS",
    "GainLossOnInvestments": "IS", "GainLossOnSaleOfBusiness": "IS",
    "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest": "IS",
    "IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments": "IS",
    "IncomeTaxExpenseBenefit": "IS",
    "NetIncomeLoss": "IS", "ProfitLoss": "IS",
    "NetIncomeLossAvailableToCommonStockholdersBasic": "IS",
    "ComprehensiveIncomeNetOfTax": "IS",
    "EarningsPerShareBasic": "IS", "EarningsPerShareDiluted": "IS",
    "WeightedAverageNumberOfSharesOutstandingBasic": "IS",
    "WeightedAverageNumberOfDilutedSharesOutstanding": "IS",
    "DepreciationDepletionAndAmortization": "IS",
    "DepreciationAndAmortization": "IS", "Depreciation": "IS",
    "ShareBasedCompensation": "IS", "AllocatedShareBasedCompensationExpense": "IS",

    # Balance Sheet
    "CashAndCashEquivalentsAtCarryingValue": "BS", "Cash": "BS",
    "CashCashEquivalentsAndShortTermInvestments": "BS",
    "RestrictedCashAndCashEquivalentsCurrent": "BS",
    "AvailableForSaleSecuritiesDebtSecuritiesCurrent": "BS",
    "MarketableSecuritiesCurrent": "BS", "ShortTermInvestments": "BS",
    "AccountsReceivableNetCurrent": "BS", "ReceivablesNetCurrent": "BS",
    "AccountsReceivableGrossCurrent": "BS",
    "InventoryNet": "BS", "InventoryGross": "BS",
    "PrepaidExpenseAndOtherAssetsCurrent": "BS", "PrepaidExpenseCurrent": "BS",
    "OtherAssetsCurrent": "BS",
    "AssetsCurrent": "BS",
    "PropertyPlantAndEquipmentNet": "BS", "PropertyPlantAndEquipmentGross": "BS",
    "AccumulatedDepreciationDepletionAndAmortizationPropertyPlantAndEquipment": "BS",
    "Goodwill": "BS",
    "IntangibleAssetsNetExcludingGoodwill": "BS", "FiniteLivedIntangibleAssetsNet": "BS",
    "AvailableForSaleSecuritiesDebtSecuritiesNoncurrent": "BS",
    "MarketableSecuritiesNoncurrent": "BS", "LongTermInvestments": "BS",
    "OperatingLeaseRightOfUseAsset": "BS",
    "DeferredIncomeTaxAssetsNet": "BS", "DeferredTaxAssetsLiabilitiesNet": "BS",
    "OtherAssetsNoncurrent": "BS",
    "Assets": "BS",
    "AccountsPayableCurrent": "BS",
    "AccruedLiabilitiesCurrent": "BS", "EmployeeRelatedLiabilitiesCurrent": "BS",
    "AccruedEmployeeBenefitsCurrent": "BS",
    "DeferredRevenueCurrent": "BS", "ContractWithCustomerLiabilityCurrent": "BS",
    "ShortTermBorrowings": "BS", "DebtCurrent": "BS", "LongTermDebtCurrent": "BS",
    "ConvertibleNotesPayableCurrent": "BS",
    "OtherLiabilitiesCurrent": "BS",
    "LiabilitiesCurrent": "BS",
    "LongTermDebt": "BS", "LongTermDebtNoncurrent": "BS",
    "ConvertibleLongTermNotesPayable": "BS", "SeniorLongTermNotes": "BS",
    "OperatingLeaseLiabilityNoncurrent": "BS",
    "DeferredRevenueNoncurrent": "BS", "ContractWithCustomerLiabilityNoncurrent": "BS",
    "DeferredIncomeTaxLiabilitiesNet": "BS",
    "OtherLiabilitiesNoncurrent": "BS",
    "Liabilities": "BS",
    "CommonStockValue": "BS",
    "AdditionalPaidInCapital": "BS", "AdditionalPaidInCapitalCommonStock": "BS",
    "RetainedEarningsAccumulatedDeficit": "BS",
    "TreasuryStockValue": "BS", "TreasuryStockCommonValue": "BS",
    "AccumulatedOtherComprehensiveIncomeLossNetOfTax": "BS",
    "StockholdersEquity": "BS",
    "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest": "BS",
    "LiabilitiesAndStockholdersEquity": "BS",
    "CommonStockSharesOutstanding": "BS",

    # Cash Flow
    "NetCashProvidedByUsedInOperatingActivities": "CF",
    "NetCashProvidedByUsedInInvestingActivities": "CF",
    "NetCashProvidedByUsedInFinancingActivities": "CF",
    "PaymentsToAcquirePropertyPlantAndEquipment": "CF",
    "PaymentsToAcquireBusinessesNetOfCashAcquired": "CF",
    "PaymentsToAcquireBusinessesGross": "CF",
    "PaymentsToAcquireAvailableForSaleSecurities": "CF",
    "PaymentsToAcquireAvailableForSaleSecuritiesDebt": "CF",
    "PaymentsToAcquireMarketableSecurities": "CF",
    "PaymentsToAcquireInvestments": "CF",
    "ProceedsFromSaleAndMaturityOfAvailableForSaleSecurities": "CF",
    "ProceedsFromSaleOfAvailableForSaleSecurities": "CF",
    "ProceedsFromMaturitiesPrepaymentsAndCallsOfAvailableForSaleSecurities": "CF",
    "ProceedsFromSaleAndMaturityOfMarketableSecurities": "CF",
    "ProceedsFromSaleMaturityAndCollectionOfInvestments": "CF",
    "ProceedsFromIssuanceOfLongTermDebt": "CF",
    "ProceedsFromConvertibleDebt": "CF", "ProceedsFromIssuanceOfDebt": "CF",
    "ProceedsFromNotesPayable": "CF", "ProceedsFromIssuanceOfSeniorLongTermDebt": "CF",
    "RepaymentsOfLongTermDebt": "CF", "RepaymentsOfConvertibleDebt": "CF",
    "RepaymentsOfDebt": "CF", "RepaymentsOfNotesPayable": "CF",
    "PaymentsForRepurchaseOfCommonStock": "CF",
    "PaymentsRelatedToTaxWithholdingForShareBasedCompensation": "CF",
    "ProceedsFromIssuanceOfCommonStock": "CF", "ProceedsFromStockOptionsExercised": "CF",
    "ProceedsFromIssuanceOfSharesUnderIncentiveAndShareBasedCompensationPlansIncludingStockOptions": "CF",
    "PaymentsOfDividends": "CF", "PaymentsOfDividendsCommonStock": "CF",
    "IncreaseDecreaseInAccountsReceivable": "CF",
    "IncreaseDecreaseInInventories": "CF",
    "IncreaseDecreaseInAccountsPayable": "CF",
    "IncreaseDecreaseInAccruedLiabilities": "CF",
    "IncreaseDecreaseInDeferredRevenue": "CF",
    "IncreaseDecreaseInOperatingCapital": "CF",
    "IncreaseDecreaseInPrepaidDeferredExpenseAndOtherAssets": "CF",
    "DeferredIncomeTaxExpenseBenefit": "CF",
    "GainLossOnDispositionOfAssets": "CF",
    "AmortizationOfFinancingCosts": "CF", "AmortizationOfDebtDiscountPremium": "CF",
    "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect": "CF",
    "CashAndCashEquivalentsPeriodIncreaseDecrease": "CF",
    "EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents": "CF",
}

# Concepts to exclude — these are footnote disclosures, not statement line items
EXCLUDE_CONCEPTS = {
    # Tax footnotes
    "IncomeTaxReconciliationIncomeTaxExpenseBenefitAtFederalStatutoryIncomeTaxRate",
    "IncomeTaxReconciliationStateAndLocalIncomeTaxes",
    "IncomeTaxReconciliationForeignIncomeTaxRateDifferential",
    "IncomeTaxReconciliationNondeductibleExpenseShareBasedCompensationCost",
    "IncomeTaxReconciliationNondeductibleExpenseResearchAndDevelopment",
    "IncomeTaxReconciliationNondeductibleExpenseOther",
    "IncomeTaxReconciliationTaxCreditsResearch",
    "IncomeTaxReconciliationChangeInEnactedTaxRate",
    "IncomeTaxReconciliationOtherAdjustments",
    "IncomeTaxReconciliationDeductions",
    "IncomeTaxReconciliationTaxContingencies",
    "IncomeTaxReconciliationFdiiAmount",
    "EffectiveIncomeTaxRateContinuingOperations",
    "EffectiveIncomeTaxRateReconciliationAtFederalStatutoryIncomeTaxRate",
    "IncomeTaxExpenseBenefitContinuingOperationsDiscontinuedOperations",
    "CurrentFederalTaxExpenseBenefit", "CurrentStateAndLocalTaxExpenseBenefit",
    "CurrentForeignTaxExpenseBenefit", "DeferredFederalIncomeTaxExpenseBenefit",
    "DeferredStateAndLocalIncomeTaxExpenseBenefit", "DeferredForeignIncomeTaxExpenseBenefit",
    "IncomeTaxesPaid", "IncomeTaxesPaidNet",
    "UnrecognizedTaxBenefits", "UnrecognizedTaxBenefitsIncomeTaxPenaltiesAndInterestAccrued",
    "UnrecognizedTaxBenefitsDecreasesResultingFromPriorPeriodTaxPositions",
    "UnrecognizedTaxBenefitsIncreasesResultingFromCurrentPeriodTaxPositions",
    "UnrecognizedTaxBenefitsIncreasesResultingFromPriorPeriodTaxPositions",
    "UnrecognizedTaxBenefitsReductionsResultingFromLapseOfApplicableStatuteOfLimitations",
    "UnrecognizedTaxBenefitsThatWouldImpactEffectiveTaxRate",
    "UndistributedEarningsOfForeignSubsidiaries",
    "IncomeTaxPaidFederalAfterRefundReceived",
    "OperatingLossCarryforwards", "DeferredTaxAssetsOperatingLossCarryforwards",
    "DeferredTaxAssetsValuationAllowance", "DeferredTaxAssetsTaxDeferredExpenseCompensationAndBenefits",
    "ValuationAllowancesAndReservesChargedToCostAndExpense",
    # Share-based comp footnotes
    "ShareBasedCompensationArrangementByShareBasedPaymentAwardOptionsGrantsInPeriodGross",
    "ShareBasedCompensationArrangementByShareBasedPaymentAwardOptionsExercisesInPeriodTotalIntrinsicValue",
    "ShareBasedCompensationArrangementByShareBasedPaymentAwardEquityInstrumentsOtherThanOptionsGrantsInPeriod",
    "EmployeeServiceShareBasedCompensationTaxBenefitFromCompensationExpense",
    "EmployeeServiceShareBasedCompensationNonvestedAwardsTotalCompensationCostNotYetRecognized",
    "ShareBasedCompensationArrangementByShareBasedPaymentAwardFairValueAssumptionsExpectedVolatilityRate",
    "ShareBasedCompensationArrangementByShareBasedPaymentAwardFairValueAssumptionsRiskFreeInterestRate",
    "AntidilutiveSecuritiesExcludedFromComputationOfEarningsPerShareAmount",
    # Debt maturity schedule footnotes
    "LongTermDebtMaturitiesRepaymentsOfPrincipalInNextTwelveMonths",
    "LongTermDebtMaturitiesRepaymentsOfPrincipalInYearTwo",
    "LongTermDebtMaturitiesRepaymentsOfPrincipalInYearThree",
    "LongTermDebtMaturitiesRepaymentsOfPrincipalInYearFour",
    "LongTermDebtMaturitiesRepaymentsOfPrincipalInYearFive",
    "LongTermDebtMaturitiesRepaymentsOfPrincipalAfterYearFive",
    "LongTermDebtMaturitiesRepaymentsOfPrincipalRemainderOfFiscalYear",
    # Fair value footnotes
    "AvailableForSaleSecuritiesAmortizedCost", "AvailableForSaleSecuritiesAccumulatedGrossUnrealizedGainBeforeTax",
    "AvailableForSaleSecuritiesAccumulatedGrossUnrealizedLossBeforeTax",
    "AvailableForSaleDebtSecuritiesAmortizedCostBasis",
    "AvailableForSaleSecuritiesDebtMaturitiesWithinOneYearAmortizedCost",
    "AvailableForSaleSecuritiesDebtMaturitiesAfterOneThroughFiveYearsAmortizedCost",
    "DebtSecuritiesAvailableForSaleAmortizedCostCurrent",
    "DebtSecuritiesAvailableForSaleAmortizedCostNoncurrent",
    "AvailableForSaleSecuritiesContinuousUnrealizedLossPosition12MonthsOrLongerFairValue",
    "AvailableForSaleSecuritiesContinuousUnrealizedLossPositionLessThan12MonthsFairValue",
    "AvailableForSaleSecuritiesGrossRealizedGains", "AvailableForSaleSecuritiesGrossRealizedLosses",
    "AvailableForSaleDebtSecuritiesAccumulatedGrossUnrealizedGainBeforeTax",
    "AvailableForSaleDebtSecuritiesAccumulatedGrossUnrealizedLossBeforeTax",
    # OCI components
    "OtherComprehensiveIncomeLossBeforeTax",
    "OtherComprehensiveIncomeLossCashFlowHedgeGainLossAfterReclassificationAndTax",
    "OtherComprehensiveIncomeLossForeignCurrencyTransactionAndTranslationAdjustmentNetOfTax",
    "OtherComprehensiveIncomeUnrealizedHoldingGainLossOnSecuritiesArisingDuringPeriodNetOfTax",
    "OtherComprehensiveIncomeLossNetOfTax", "OtherComprehensiveIncomeUnrealizedHoldingGainLossOnSecuritiesArisingDuringPeriodBeforeTax",
    "OtherComprehensiveIncomeLossForeignCurrencyTranslationAdjustmentTax",
    # Entity-level / DEI
    "EntityCommonStockSharesOutstanding", "EntityPublicFloat", "EntityNumberOfEmployees",
    "DocumentFiscalYearFocus", "DocumentFiscalPeriodFocus",
    "CommonStockSharesAuthorized", "CommonStockParOrStatedValuePerShare",
    "CommonStockSharesIssued", "PreferredStockSharesAuthorized",
    "PreferredStockSharesIssued", "PreferredStockSharesOutstanding",
    # Lease footnotes
    "OperatingLeaseWeightedAverageRemainingLeaseTerm1",
    "OperatingLeaseWeightedAverageDiscountRatePercent",
    "LesseeOperatingLeaseLiabilityPaymentsDue",
    "LesseeOperatingLeaseLiabilityPaymentsDueNextTwelveMonths",
    "LesseeOperatingLeaseLiabilityPaymentsDueYearTwo",
    "LesseeOperatingLeaseLiabilityPaymentsDueYearThree",
    "LesseeOperatingLeaseLiabilityPaymentsDueYearFour",
    "LesseeOperatingLeaseLiabilityPaymentsDueYearFive",
    "OperatingLeaseExpense", "OperatingLeaseCost",
    "OperatingLeasesRentExpenseNet", "ShortTermLeaseCost",
    "OperatingLeaseImpairmentLoss", "OperatingLeaseRightOfUseAssetAmortizationExpense",
    # Misc footnotes
    "DefinedContributionPlanEmployerDiscretionaryContributionAmount",
    "DefinedContributionPlanEmployerMatchingContributionPercent",
    "CumulativeEffectOfNewAccountingPrincipleInPeriodOfAdoption",
    "OtherRestructuringCosts", "RestructuringReserveCurrent",
    "InterestCostsCapitalized", "InterestPaidNet", "InterestPaid",
    "CapitalExpendituresIncurredButNotYetPaid",
    "CommonStockCapitalSharesReservedForFutureIssuance",
    "DebtInstrumentPeriodicPaymentInterest",
    "SeveranceCosts1", "BusinessCombinationIntegrationRelatedCosts",
    "ForeignCurrencyTransactionGainLossUnrealized",
    "PaidInKindInterest", "OtherThanTemporaryImpairmentLossesInvestments",
    "IncomeTaxReconciliationIncomeTaxExpenseBenefitAtFederalStatutoryIncomeTaxRate",
    "AccretionAmortizationOfDiscountsAndPremiumsInvestments",
    "IncreaseDecreaseInDeferredIncomeTaxes",
}

# Concepts reported as USD/shares (EPS) — don't divide by 1MM
PER_SHARE_CONCEPTS = {
    "EarningsPerShareBasic", "EarningsPerShareDiluted",
    "EarningsPerShareBasicAndDiluted",
}

# Concepts reported as share counts — convert to MM shares
SHARE_COUNT_CONCEPTS = {
    "WeightedAverageNumberOfSharesOutstandingBasic",
    "WeightedAverageNumberOfDilutedSharesOutstanding",
    "CommonStockSharesOutstanding",
}


def human_label(concept: str) -> str:
    """CamelCase XBRL name → readable label."""
    s = re.sub(r"([A-Z][a-z]+)", r" \1", re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1 \2", concept))
    return s.strip()


def extract_concepts(facts: dict, is_annual: bool, cutoff_date: str) -> dict[str, dict]:
    """
    Pull every non-footnote financial concept, assign to IS/BS/CF,
    convert to $MM, and return {concept: {period: value}}.
    """
    valid_forms = ("10-K", "20-F") if is_annual else ("10-Q", "6-K")
    # Annual = ~12 month window, quarterly = ~3 month window
    span_min, span_max = (300, 400) if is_annual else (60, 110)
    # Balance sheet items are instantaneous (no start date) — allow those through
    bs_concepts = {c for c, s in CONCEPT_STATEMENT.items() if s == "BS"}

    result: dict[str, dict] = {}

    for ns in ["us-gaap", "dei"]:
        ns_facts = facts.get(ns, {})
        for concept, concept_data in ns_facts.items():
            if concept in EXCLUDE_CONCEPTS:
                continue

            # Only include concepts we've explicitly classified, OR
            # concepts that look like they belong on a statement
            stmt = CONCEPT_STATEMENT.get(concept)
            if stmt is None:
                continue  # skip unknown/footnote concepts entirely

            units = concept_data.get("units", {})
            if "USD" in units:
                raw, unit_type = units["USD"], "USD"
            elif "USD/shares" in units:
                raw, unit_type = units["USD/shares"], "USD/shares"
            elif "shares" in units:
                raw, unit_type = units["shares"], "shares"
            else:
                continue

            filtered = [
                e for e in raw
                if e.get("form") in valid_forms
                and e.get("end", "") >= cutoff_date
                and e.get("val") is not None
            ]
            if not filtered:
                continue

            period_map: dict[str, dict] = {}
            for e in filtered:
                end   = e["end"]
                filed = e.get("filed", "")
                val   = e["val"]
                start = e.get("start", "")

                # Span filter — skip for BS (instantaneous) and for concepts
                # without a start date
                is_bs = concept in bs_concepts
                if not is_bs and start:
                    try:
                        span = (pd.Timestamp(end) - pd.Timestamp(start)).days
                        if not (span_min <= span <= span_max):
                            continue
                    except Exception:
                        pass

                period_key = end[:4] if is_annual else end

                if period_key not in period_map or filed > period_map[period_key]["filed"]:
                    period_map[period_key] = {"val": val, "filed": filed}

            if not period_map:
                continue

            # Unit conversion
            converted = {}
            for pk, entry in period_map.items():
                v = entry["val"]
                if concept in PER_SHARE_CONCEPTS or unit_type == "USD/shares":
                    converted[pk] = round(v, 4)
                elif concept in SHARE_COUNT_CONCEPTS or unit_type == "shares":
                    converted[pk] = round(v / 1_000_000, 3)
                else:  # USD → $MM
                    converted[pk] = round(v / 1_000_000, 3)

            label = human_label(concept)
            key   = (stmt, label)
            if key not in result or ns == "us-gaap":
                result[key] = converted

    # Reshape to {stmt: {label: {period: val}}}
    by_stmt: dict[str, dict] = {"IS": {}, "BS": {}, "CF": {}}
    for (stmt, label), period_data in result.items():
        by_stmt[stmt][label] = period_data

    return by_stmt


def get_sorted_periods(by_stmt: dict, is_annual: bool) -> list[str]:
    all_periods: set[str] = set()
    for stmt_data in by_stmt.values():
        for period_data in stmt_data.values():
            all_periods.update(period_data.keys())
    return sorted(all_periods, reverse=True)


# ── Preferred ordering within each statement ───────────────────────────────────

IS_ORDER = [
    "Revenues", "Revenue From Contract With Customer Excluding Assessed Tax",
    "Revenue From Contract With Customer Including Assessed Tax",
    "Sales Revenue Net", "Sales Revenue Goods Net", "Royalty Revenue",
    "License And Services Revenue", "Licenses Revenue",
    "Cost Of Revenue", "Cost Of Goods Sold", "Cost Of Goods And Services Sold", "Cost Of Services",
    "Gross Profit",
    "Research And Development Expense", "Research And Development Expense Excluding Acquired In Process Cost",
    "Selling General And Administrative Expense", "General And Administrative Expense",
    "Selling And Marketing Expense", "Selling Expense",
    "Amortization Of Intangible Assets", "Amortization Of Acquired Intangible Assets",
    "Goodwill Impairment Loss", "Impairment Of Intangible Assets Excluding Goodwill",
    "Restructuring Charges", "Business Combination Acquisition Related Costs",
    "Operating Expenses", "Costs And Expenses",
    "Operating Income Loss",
    "Interest Expense", "Interest And Debt Expense",
    "Investment Income Nonoperating", "Investment Income Interest",
    "Interest Income Expense Net", "Interest Income Expense Nonoperating Net",
    "Nonoperating Income Expense", "Other Nonoperating Income Expense",
    "Income Loss From Continuing Operations Before Income Taxes Extraordinary Items Noncontrolling Interest",
    "Income Tax Expense Benefit",
    "Net Income Loss", "Profit Loss", "Net Income Loss Available To Common Stockholders Basic",
    "Comprehensive Income Net Of Tax",
    "Earnings Per Share Basic", "Earnings Per Share Diluted",
    "Weighted Average Number Of Shares Outstanding Basic",
    "Weighted Average Number Of Diluted Shares Outstanding",
    "Depreciation Depletion And Amortization", "Depreciation And Amortization",
    "Share Based Compensation", "Allocated Share Based Compensation Expense",
]

BS_ORDER = [
    "Cash And Cash Equivalents At Carrying Value",
    "Available For Sale Securities Debt Securities Current", "Marketable Securities Current", "Short Term Investments",
    "Accounts Receivable Net Current", "Receivables Net Current",
    "Inventory Net",
    "Prepaid Expense And Other Assets Current", "Other Assets Current",
    "Assets Current",
    "Property Plant And Equipment Net",
    "Goodwill",
    "Intangible Assets Net Excluding Goodwill", "Finite Lived Intangible Assets Net",
    "Available For Sale Securities Debt Securities Noncurrent", "Long Term Investments",
    "Operating Lease Right Of Use Asset",
    "Deferred Income Tax Assets Net",
    "Other Assets Noncurrent",
    "Assets",
    "Accounts Payable Current",
    "Accrued Liabilities Current", "Employee Related Liabilities Current",
    "Deferred Revenue Current", "Contract With Customer Liability Current",
    "Debt Current", "Short Term Borrowings", "Long Term Debt Current", "Convertible Notes Payable Current",
    "Other Liabilities Current",
    "Liabilities Current",
    "Long Term Debt Noncurrent", "Long Term Debt", "Convertible Long Term Notes Payable",
    "Operating Lease Liability Noncurrent",
    "Deferred Revenue Noncurrent", "Contract With Customer Liability Noncurrent",
    "Deferred Income Tax Liabilities Net",
    "Other Liabilities Noncurrent",
    "Liabilities",
    "Additional Paid In Capital", "Additional Paid In Capital Common Stock",
    "Retained Earnings Accumulated Deficit",
    "Treasury Stock Value", "Treasury Stock Common Value",
    "Accumulated Other Comprehensive Income Loss Net Of Tax",
    "Stockholders Equity",
    "Liabilities And Stockholders Equity",
    "Common Stock Shares Outstanding",
]

CF_ORDER = [
    "Net Income Loss", "Profit Loss",
    "Depreciation Depletion And Amortization", "Depreciation And Amortization",
    "Amortization Of Intangible Assets",
    "Share Based Compensation", "Allocated Share Based Compensation Expense",
    "Deferred Income Tax Expense Benefit",
    "Amortization Of Financing Costs", "Amortization Of Debt Discount Premium",
    "Increase Decrease In Accounts Receivable",
    "Increase Decrease In Inventories",
    "Increase Decrease In Prepaid Deferred Expense And Other Assets",
    "Increase Decrease In Accounts Payable",
    "Increase Decrease In Accrued Liabilities",
    "Increase Decrease In Deferred Revenue",
    "Increase Decrease In Operating Capital",
    "Net Cash Provided By Used In Operating Activities",
    "Payments To Acquire Property Plant And Equipment",
    "Payments To Acquire Businesses Net Of Cash Acquired",
    "Payments To Acquire Available For Sale Securities",
    "Payments To Acquire Available For Sale Securities Debt",
    "Payments To Acquire Marketable Securities",
    "Proceeds From Sale And Maturity Of Available For Sale Securities",
    "Proceeds From Sale Of Available For Sale Securities",
    "Proceeds From Maturities Prepayments And Calls Of Available For Sale Securities",
    "Proceeds From Sale And Maturity Of Marketable Securities",
    "Net Cash Provided By Used In Investing Activities",
    "Proceeds From Issuance Of Long Term Debt",
    "Proceeds From Convertible Debt", "Proceeds From Issuance Of Debt",
    "Repayments Of Long Term Debt", "Repayments Of Convertible Debt", "Repayments Of Debt",
    "Payments For Repurchase Of Common Stock",
    "Payments Related To Tax Withholding For Share Based Compensation",
    "Proceeds From Issuance Of Common Stock", "Proceeds From Stock Options Exercised",
    "Payments Of Dividends", "Payments Of Dividends Common Stock",
    "Net Cash Provided By Used In Financing Activities",
    "Cash Cash Equivalents Restricted Cash And Restricted Cash Equivalents Period Increase Decrease Including Exchange Rate Effect",
    "Cash And Cash Equivalents Period Increase Decrease",
]

def sort_statement(data: dict, order: list[str]) -> dict:
    """Sort statement line items: preferred order first, then alphabetical for remainder."""
    order_index = {label: i for i, label in enumerate(order)}
    def sort_key(label):
        return (order_index.get(label, len(order)), label)
    return dict(sorted(data.items(), key=lambda x: sort_key(x[0])))


# ── Excel builder ──────────────────────────────────────────────────────────────

HDR_FILL   = PatternFill("solid", start_color="1F3864")
ALT_FILL   = PatternFill("solid", start_color="EBF3FB")
WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
LABEL_FONT = Font(name="Arial", size=9)
BOLD_FONT  = Font(name="Arial", bold=True, size=9)
DATA_FONT  = Font(name="Arial", size=9, color="00008B")

TOTAL_LABELS = {
    "Gross Profit", "Operating Income Loss", "Assets Current", "Liabilities Current",
    "Assets", "Liabilities", "Stockholders Equity", "Liabilities And Stockholders Equity",
    "Net Income Loss", "Profit Loss", "Comprehensive Income Net Of Tax",
    "Net Cash Provided By Used In Operating Activities",
    "Net Cash Provided By Used In Investing Activities",
    "Net Cash Provided By Used In Financing Activities",
    "Costs And Expenses", "Operating Expenses",
}

def write_statement_sheet(wb, sheet_name, data, periods, company_name, period_label):
    if not data or not periods:
        return
    ws = wb.create_sheet(sheet_name[:31])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"

    # Title row
    ws.row_dimensions[1].height = 20
    ws.cell(row=1, column=1).value = f"{company_name}  —  {sheet_name}"
    ws.cell(row=1, column=1).font  = Font(name="Arial", bold=True, size=11, color="1F3864")
    nc = len(periods)
    ws.cell(row=1, column=nc+2).value = "$MM where applicable  |  EPS in $/share  |  Shares in MM  |  Source: SEC EDGAR XBRL (as reported)"
    ws.cell(row=1, column=nc+2).font  = Font(name="Arial", size=8, color="888888", italic=True)

    # Header row
    ws.row_dimensions[2].height = 16
    ws.cell(row=2, column=1).value = period_label
    ws.cell(row=2, column=1).font  = HDR_FONT
    ws.cell(row=2, column=1).fill  = HDR_FILL
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for i, p in enumerate(periods):
        c = i + 2
        ws.cell(row=2, column=c).value     = p
        ws.cell(row=2, column=c).font      = HDR_FONT
        ws.cell(row=2, column=c).fill      = HDR_FILL
        ws.cell(row=2, column=c).alignment = Alignment(horizontal="center", vertical="center")

    for row_i, (label, period_data) in enumerate(data.items()):
        r = row_i + 3
        ws.row_dimensions[r].height = 14
        fill = ALT_FILL if row_i % 2 == 0 else WHITE_FILL
        is_total = label in TOTAL_LABELS

        ws.cell(row=r, column=1).value     = label
        ws.cell(row=r, column=1).font      = BOLD_FONT if is_total else LABEL_FONT
        ws.cell(row=r, column=1).fill      = fill
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=1)

        for j, p in enumerate(periods):
            c    = j + 2
            val  = period_data.get(p)
            cell = ws.cell(row=r, column=c)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="right")
            if val is not None:
                cell.value = val
                cell.font  = BOLD_FONT if is_total else DATA_FONT
                cell.number_format = ('#,##0.00;(#,##0.00);"-"' if abs(val) < 100 and val != int(val)
                                      else '#,##0.0;(#,##0.0);"-"')
            else:
                cell.value = "—"
                cell.font  = Font(name="Arial", size=9, color="BBBBBB")

    ws.column_dimensions["A"].width = 56
    for i in range(len(periods)):
        ws.column_dimensions[get_column_letter(i+2)].width = 13


def write_raw_table_sheet(wb, sheet_name, df, title):
    ws = wb.create_sheet(sheet_name[:31])
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1).value = title
    ws.cell(row=1, column=1).font  = Font(name="Arial", bold=True, size=10, color="1F3864")
    for c_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=c_idx)
        cell.value = str(col); cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
    for r_idx, row in df.iterrows():
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx+3, column=c_idx)
            cell.value = val; cell.font = Font(name="Arial", size=9)
            cell.fill = ALT_FILL if r_idx % 2 == 0 else WHITE_FILL
    for c_idx in range(1, len(df.columns)+1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 24


# ── Segment / KPI HTML extraction ─────────────────────────────────────────────

JUNK_PATTERN = re.compile(
    r"table of contents|exhibit index|incorporated herein by reference|"
    r"certification of chief|pursuant to rule 13a|pursuant to section 906|"
    r"instance document|taxonomy extension|inline xbrl|"
    r"ernst.*young|deloitte|kpmg|pricewaterhousecoopers|/s/ |"
    r"trading arrangement|rule 10b5|shares to be sold|expiration date|"
    r"accounting standard|fasb issued|asc 842|adoption method|"
    r"bylaws|certificate of incorporation|indenture.*trustee",
    re.IGNORECASE
)

def is_useful_table(df):
    if df.shape[0] < 3 or df.shape[1] < 2:
        return False
    sample = " ".join(str(v) for v in df.iloc[:4].values.flatten() if v)
    if JUNK_PATTERN.search(sample):
        return False
    all_text = " ".join(str(v) for v in df.values.flatten() if v)
    if not re.search(r"\d{2,}", all_text):
        return False
    non_empty = sum(1 for v in df.values.flatten() if str(v).strip() not in ("", "nan", "None"))
    return non_empty / max(df.size, 1) >= 0.15

def fetch_filing_html(cik, accession, primary_doc):
    url = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{accession}/{primary_doc}"
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        if r.status_code == 200:
            return r.text
    except Exception:
        pass
    return None

def extract_tables_from_html(html, keywords):
    soup = BeautifulSoup(html, "html.parser")
    results = []; seen = set()
    for kw in keywords:
        pattern = re.compile(r"\b" + re.escape(kw) + r"\b", re.IGNORECASE)
        for match in soup.find_all(string=pattern)[:5]:
            parent = match.parent
            for _ in range(8):
                if parent is None: break
                table = parent.find_next("table")
                if table:
                    try:
                        dfs = pd.read_html(str(table))
                        if dfs:
                            df = dfs[0]
                            fp = str(df.values.tolist())[:300]
                            if fp not in seen and is_useful_table(df):
                                seen.add(fp); results.append(df)
                    except Exception:
                        pass
                    break
                parent = parent.parent
    return results

def fetch_segment_data(cik, filings, max_filings=8):
    keywords = [
        "segment revenue", "segment information", "revenue by segment",
        "geographic", "revenue by region", "revenue by geography",
        "disaggregated revenue", "revenue disaggregation",
        "customer concentration", "significant customer", "major customer",
        "royalt", "product sales", "collaborative",
        "subscribers", "active users", "monthly active", "annual recurring",
        "units sold", "same-store", "comparable store",
        "backlog", "bookings", "net revenue retention",
        "key performance", "operating metric",
    ]
    results = []; seen_tables: set = set()
    for filing in filings[:max_filings]:
        html = fetch_filing_html(cik, filing["accession"], filing["primary_doc"])
        if not html: continue
        for df in extract_tables_from_html(html, keywords):
            key = str(df.values.tolist())[:300]
            if key not in seen_tables:
                seen_tables.add(key)
                results.append((filing["date"], df.dropna(how="all").fillna("")))
        time.sleep(0.3)
    return results


# ── Streamlit UI ───────────────────────────────────────────────────────────────

def main():
    st.title("📊 SEC Financial Data Extractor")
    st.markdown("Pulls financial statement line items **as reported** from SEC EDGAR XBRL filings.")

    with st.sidebar:
        st.header("Settings")
        ticker  = st.text_input("Stock Ticker", placeholder="e.g. AAPL, MSFT, HALO").strip().upper()
        years   = st.slider("Years of History", 5, 20, 15)
        include_segments = st.checkbox("Extract Segment / KPI Tables (slower)", value=True)
        max_seg = st.slider("# Filings to Scan for Segments", 4, 20, 8)
        run_btn = st.button("🚀 Extract Data", type="primary", use_container_width=True)

    if not run_btn or not ticker:
        st.info("Enter a ticker in the sidebar and click **Extract Data** to begin.")
        st.markdown("""
**What this tool extracts:**
- ✅ Income Statement line items as reported (no aggregation)
- ✅ Balance Sheet as reported
- ✅ Cash Flow Statement as reported
- ✅ Quarterly and Annual tabs for each statement
- ✅ Segment / geographic / KPI tables (best-effort from filing HTML)
- ✅ All dollar values in $MM · EPS in $/share · Shares in MM
        """)
        return

    progress = st.progress(0); status = st.empty()

    try:
        status.text("🔍 Looking up CIK...")
        cik = get_cik(ticker)
        if not cik:
            st.error(f"Could not find CIK for '{ticker}'.")
            return
        progress.progress(5)

        company_name = get_company_name(cik)
        st.success(f"Found: **{company_name}** (CIK: {int(cik)})")

        status.text("📥 Downloading XBRL company facts...")
        facts = get_xbrl_facts(cik)
        progress.progress(20)

        cutoff_a = f"{pd.Timestamp.now().year - years}-01-01"
        cutoff_q = str((pd.Timestamp.now() - pd.DateOffset(years=years)).date())

        status.text("🔢 Extracting annual data...")
        annual = extract_concepts(facts, is_annual=True,  cutoff_date=cutoff_a)
        annual_periods = list(reversed(get_sorted_periods(annual, True)))
        progress.progress(40)

        status.text("🔢 Extracting quarterly data...")
        qtr = extract_concepts(facts, is_annual=False, cutoff_date=cutoff_q)
        qtr_periods = get_sorted_periods(qtr, False)[:60]
        progress.progress(60)

        # Sort each statement in logical order
        annual["IS"] = sort_statement(annual["IS"], IS_ORDER)
        annual["BS"] = sort_statement(annual["BS"], BS_ORDER)
        annual["CF"] = sort_statement(annual["CF"], CF_ORDER)
        qtr["IS"]    = sort_statement(qtr["IS"],    IS_ORDER)
        qtr["BS"]    = sort_statement(qtr["BS"],    BS_ORDER)
        qtr["CF"]    = sort_statement(qtr["CF"],    CF_ORDER)

        segment_tables = []
        if include_segments:
            status.text("🔍 Fetching filings for segment/KPI data...")
            all_filings = sorted(
                get_filings_index(cik, ["10-K", "10-Q"]),
                key=lambda x: x["date"], reverse=True
            )
            segment_tables = fetch_segment_data(cik, all_filings, max_seg)
        progress.progress(80)

        status.text("📝 Building Excel workbook...")
        wb = Workbook(); wb.remove(wb.active)

        write_statement_sheet(wb, "Annual — Income Stmt",   annual["IS"], annual_periods, company_name, "Fiscal Year")
        write_statement_sheet(wb, "Annual — Balance Sheet", annual["BS"], annual_periods, company_name, "Fiscal Year")
        write_statement_sheet(wb, "Annual — Cash Flow",     annual["CF"], annual_periods, company_name, "Fiscal Year")
        write_statement_sheet(wb, "Qtrly — Income Stmt",    qtr["IS"],    qtr_periods,    company_name, "Quarter Ended")
        write_statement_sheet(wb, "Qtrly — Balance Sheet",  qtr["BS"],    qtr_periods,    company_name, "Quarter Ended")
        write_statement_sheet(wb, "Qtrly — Cash Flow",      qtr["CF"],    qtr_periods,    company_name, "Quarter Ended")

        for idx, (filing_date, df) in enumerate(segment_tables[:25]):
            write_raw_table_sheet(wb, f"Seg-KPI {filing_date} ({idx+1})", df,
                                  f"Extracted Table — Filing: {filing_date}")

        progress.progress(95)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        progress.progress(100); status.text("✅ Done!")

        fname = f"{ticker}_financials_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx"
        st.download_button("⬇️ Download Excel File", data=buf, file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           type="primary")

        # Preview
        st.subheader("Preview — Annual Income Statement")
        if annual["IS"]:
            preview = [{"Line Item": lbl, **{p: pd.get(p, "—") for p in annual_periods[:10]}}
                       for lbl, pd in annual["IS"].items()]
            st.dataframe(pd_df := pd.DataFrame(preview).set_index("Line Item"), use_container_width=True)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Annual IS lines",     len(annual["IS"]))
        c2.metric("Annual BS lines",     len(annual["BS"]))
        c3.metric("Annual CF lines",     len(annual["CF"]))
        c4.metric("Seg/KPI tables",      len(segment_tables))

        if segment_tables:
            with st.expander(f"Preview first Seg/KPI table ({segment_tables[0][0]})"):
                st.dataframe(segment_tables[0][1], use_container_width=True)

    except Exception as e:
        st.error(f"Error: {e}")
        st.exception(e)

if __name__ == "__main__":
    main()
